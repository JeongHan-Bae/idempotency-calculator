from orbit import *
import openpyxl
from openpyxl.styles import *
import json

with open('code.json', 'r') as f:
    obj = json.load(f)
    code = obj['code']
    debut = obj['debut']
    f.close()
workbook = openpyxl.load_workbook("idem_input.xlsx")
worksheet = workbook.active
worksheet['A1'] = "ASCII"
worksheet['J1'] = "Notre Pondération"
worksheet['T1'] = "Faro_in"
worksheet['U1'] = "Faro_out"
worksheet['V1'] = "Monge_in"
worksheet['W1'] = "Monge_out"
for i in range(1, 258):
    for j in range(0, 8):
        worksheet.cell(i, 11 + j).value = worksheet.cell(i, 9 - code[j]).value
for i in range(11, 19):
    worksheet.cell(1, i).value = worksheet.cell(1, i - 9).value + "(" + worksheet.cell(1, i).value + ")"

for i in range(2, 258):
    info = []
    if debut == "e0":
        for j in range(18, 10, -1):
            info.append(worksheet.cell(i, j).value)
    else:
        for j in range(11, 19):
            info.append(worksheet.cell(i, j).value)
    worksheet.cell(i, 20).value = idempotence(info, faro_in(8))
    worksheet.cell(i, 21).value = idempotence(info, faro_out(8))
    worksheet.cell(i, 22).value = idempotence(info, monge_in(8))
    worksheet.cell(i, 23).value = idempotence(info, monge_out(8))

worksheet.column_dimensions['J'].width = 20
worksheet.column_dimensions['T'].width = 15
worksheet.column_dimensions['U'].width = 15
worksheet.column_dimensions['V'].width = 15
worksheet.column_dimensions['W'].width = 15

worksheet['A1'].font = Font(color='FFFFFF', bold=True)
worksheet['A1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

worksheet['J1'].font = Font(color='000000', bold=True)
worksheet['J1'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

for row in worksheet.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
workbook.save("idem_output.xlsx")
